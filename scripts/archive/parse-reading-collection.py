#!/usr/bin/env python3
"""
Parse Reading Collection.xlsx and output to JSON.
Maps topics to RSTU categories and formats for metadata matching.
"""

import json
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --quiet")
    import openpyxl

XLSX_FILE = Path.home() / "Documents/References/Reading Collection.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "data/reading-collection.json"

# Map topics to RSTU categories
TOPIC_TO_CATEGORY = {
    "Housing": "housing-rent-tenants",
    "Politics": "organizing-action",
    "History": "historical",
    "Economics": "classic-theory",
    "Philosophy": "classic-theory",
    "Sociology": "contemporary-analysis",
    "Labor": "labor-unions",
    "Organizing": "organizing-action",
    "Police": "police-enforcement",
    "Prison": "police-enforcement",
    "Environment": "contemporary-analysis",
    "Education": "contemporary-analysis",
}

def normalize_title(title):
    """Normalize title for matching."""
    if not title:
        return ""
    return str(title).lower().strip()

def parse_xlsx():
    if not XLSX_FILE.exists():
        print(f"❌ File not found: {XLSX_FILE}")
        return None

    print(f"📖 Loading: {XLSX_FILE}")
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"📋 Columns: {headers}")

    documents = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue

        title = str(row[0]).strip() if row[0] else None
        author = str(row[1]).strip() if row[1] else None
        # Handle non-numeric values
        try:
            publication = int(float(row[2])) if row[2] and str(row[2]).strip() not in ['-', ''] else None
        except (ValueError, TypeError):
            publication = None
        try:
            page_count = int(float(row[3])) if row[3] and str(row[3]).strip() not in ['-', ''] else None
        except (ValueError, TypeError):
            page_count = None
        topic = str(row[4]).strip() if row[4] else None
        status = str(row[5]).strip() if row[5] else None

        if not title:
            continue

        # Map topic to category
        category = TOPIC_TO_CATEGORY.get(topic, "contemporary-analysis")

        documents.append({
            "title": title,
            "normalized_title": normalize_title(title),
            "author": author,
            "year": publication,
            "page_count": page_count,
            "topic": topic,
            "category": category,
            "status": status,
        })

    wb.close()
    return documents

def main():
    print("📚 Reading Collection Parser")
    print("=" * 50)

    documents = parse_xlsx()
    if not documents:
        return

    print(f"✅ Parsed {len(documents)} documents")

    # Statistics
    with_author = sum(1 for d in documents if d["author"])
    with_year = sum(1 for d in documents if d["year"])
    topics = {}
    for d in documents:
        t = d["topic"] or "Unknown"
        topics[t] = topics.get(t, 0) + 1

    print(f"\n📊 Statistics:")
    print(f"   With author: {with_author} ({with_author/len(documents)*100:.1f}%)")
    print(f"   With year: {with_year} ({with_year/len(documents)*100:.1f}%)")
    print(f"\n📁 Topics:")
    for topic, count in sorted(topics.items(), key=lambda x: -x[1]):
        print(f"   {topic}: {count}")

    # Ensure output directory exists
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # Write output
    output = {
        "generated": __import__("datetime").datetime.now().isoformat(),
        "source": str(XLSX_FILE),
        "total": len(documents),
        "statistics": {
            "with_author": with_author,
            "with_year": with_year,
            "topics": topics,
        },
        "documents": documents,
    }

    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\n📁 Saved to: {OUTPUT_FILE}")

    # Show sample
    print("\n📋 Sample documents:")
    for d in documents[:5]:
        print(f"   • {d['title'][:50]}...")
        print(f"     Author: {d['author']}, Year: {d['year']}, Topic: {d['topic']}")

if __name__ == "__main__":
    main()
